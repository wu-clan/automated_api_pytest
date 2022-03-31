#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import shutil
from datetime import datetime
from typing import Optional

import openpyxl
import xlrd
from openpyxl.styles import Font, Alignment, Color
from openpyxl.styles.colors import COLOR_INDEX

from fastpt.common.log import log
from fastpt.core.get_conf import TESTER_NAME
from fastpt.core.path_conf import EXCEL_DATA_PATH, EXCEL_REPORT_PATH, EXCEL_REPORT_TEMPLATE_FILE


def read_excel(filepath: str = EXCEL_DATA_PATH, *, filename: str, sheet: str = 'Sheet1', ) -> Optional[list]:
    """
    读取 xlsx 文件
    :param filepath: 文件路径
    :param filename: 文件名
    :param sheet: 工作表
    :return:
    """
    file = os.path.join(filepath, filename)
    data = xlrd.open_workbook(file)
    table = data.sheet_by_name(sheet)
    # 获取总行,列数
    rows = table.nrows
    cols = table.ncols
    if rows > 1:
        # 获取第一行内容,通常为列说明
        keys = table.row_values(0)
        data_list = []
        # 获取文档剩下所有内容
        for col in range(1, cols):
            values = table.row_values(col)
            # key, value组合为字典
            data = dict(zip(keys, values))
            data_list.append(data)
        return data_list
    else:
        log.warning('数据表格没有数据!')
        return None


def write_excel_report(datafile='APITestCaseTEMPLATE.xlsx', filename: str = f'APITestResult_{datetime.now()}.xlsx', *,
                       row_n: int, status: str):
    """
    写入excel测试报告
    :param datafile: excel测试数据文件名
    :param filename: 文件名
    :param row_n:数据所在行数
    :param status: 测试结果: 'PASS' or 'FAIL'
    :return
    """
    if not os.path.exists(EXCEL_REPORT_PATH):
        os.makedirs(EXCEL_REPORT_PATH)
    _datafile = os.path.join(EXCEL_DATA_PATH, datafile)
    _report_file = os.path.join(EXCEL_REPORT_PATH, filename)
    # copy测试数据为报告文件基础
    shutil.copyfile(datafile, _report_file)
    wb = openpyxl.load_workbook(_report_file)
    ws = wb.active
    font_green = Font(name='宋体', color=Color(rgb=COLOR_INDEX[3]), bold=True)
    font_red = Font(name='宋体', color=Color(rgb=COLOR_INDEX[2]), bold=True)
    font_black = Font(name='宋体', color=Color(), bold=True)
    # 文件内容格式
    align = Alignment(horizontal='center', vertical='center')
    # 所在行,列
    L_n = "L" + str(row_n)
    M_n = "M" + str(row_n)
    if status == "PASS":
        ws.cell(row_n, 12, status)
        ws[L_n].font = font_green
    if status == "FAIL":
        ws.cell(row_n, 12, status)
        ws[L_n].font = font_red
    ws.cell(row_n, 13, TESTER_NAME)
    ws[M_n].font = font_black
    ws[L_n].alignment = ws[M_n].alignment = align
    try:
        wb.save(_report_file)
    except Exception as e:
        log.error(f'写入excel测试报告失败 \n {e}')
        raise e
    else:
        log.info('写入excel测试报告成功')